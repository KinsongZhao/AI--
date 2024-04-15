import json
import re

import openpyxl
from openpyxl.styles import Alignment
from playhouse.shortcuts import model_to_dict

from dianping.db.local_db import Store

if __name__ == '__main__':
    file_path = './report_dianping.xlsx'
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    max_row = sheet.max_row

    sheet.column_dimensions['D'].width = 100

    list = Store.select()
    for key, item in enumerate(list):
        max_row = sheet.max_row
        store_model = model_to_dict(item)
        sheet.cell(max_row + 1, 1).value = item.shopName
        sheet.cell(max_row + 1, 2).value = item.addr
        sheet.cell(max_row + 1, 3).value = item.originalPrice
        posdescr = ''
        if item.posdescr is not None:
            desc_list = json.loads(item.posdescr)
            for key, it in enumerate(desc_list):
                if key >= 3:
                    break
                posdescr += it['content'] + "|"
        sheet.cell(max_row + 1, 4).value = posdescr
        # 创建一个Alignment对象，并设置wrap_text属性为True
        alignment = Alignment(wrap_text=True)
        sheet[f'D{key + 1}'].alignment = alignment
        if hasattr(item, 'areaName'):
            sheet.cell(max_row + 1, 5).value = item.areaName


        # 定义正则表达式
        regex = re.compile(r"\d+")
        # 使用findall方法获取所有匹配的浮点数
        originalPrice = re.findall(regex, item.originalPrice)
        if len(originalPrice) > 0:
            originalPrice = originalPrice[0]
            xiaofei = ''
            if float(originalPrice) < 100:
                xiaofei = '低消费'
            if float(originalPrice) >= 100 and float(originalPrice) <= 300:
                xiaofei = '中等消费'
            if float(originalPrice) > 300:
                xiaofei = '低消费'
            sheet.cell(max_row + 1, 6).value =xiaofei
        sheet.cell(max_row + 1, 7).value = item.hotelStar

    workbook.save(file_path)
